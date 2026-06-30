import importlib.util
import json
import re
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Volumes/盛田協作/Ryuk/國三落點分析")
SITE = ROOT / "school-map"
INPUT_XLSX = ROOT / "outputs/high_school_lowest3/大華國中104-114高中職各校科最低三筆_學校名稱整理版.xlsx"
CURRENT_JSON = SITE / "data/schools.json"
ADDRESS_XLSX = ROOT / "outputs/extracted_school_data/115中投區_29-40頁_學校地址科系.xlsx"
ARCGIS_CACHE = SITE / "data/arcgis-address-cache.json"
OUT_JSON = SITE / "data/schools.json"

STANDARDIZER_PATH = ROOT / "outputs/high_school_lowest3/work/extract_high_school_lowest3.py"

spec = importlib.util.spec_from_file_location("lowest3_extract", STANDARDIZER_PATH)
lowest3_extract = importlib.util.module_from_spec(spec)
spec.loader.exec_module(lowest3_extract)
standardize_school_name = lowest3_extract.standardize_school_name


def text(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_address(address):
    address = text(address)
    address = re.sub(r"^\(\d+\)", "", address)
    return address.strip()


def address_key(address):
    return clean_address(address).replace("臺", "台")


def grade_group(grade):
    grade = text(grade)
    if grade.startswith("A"):
        return "A"
    if grade.startswith("B"):
        return "B"
    if grade.startswith("C"):
        return "C"
    return ""


def recalc_school(school):
    records = [record for dept in school["departments"] for record in dept["records"]]
    points = [record["totalPoints"] for record in records if isinstance(record.get("totalPoints"), (int, float))]
    school["recordCount"] = len(records)
    school["avgTotalPoints"] = round(sum(points) / len(points), 1) if points else None
    school["minTotalPoints"] = min(points) if points else None
    school["maxTotalPoints"] = max(points) if points else None
    for dept in school["departments"]:
        dept_points = [record["totalPoints"] for record in dept["records"] if isinstance(record.get("totalPoints"), (int, float))]
        dept["avgTotalPoints"] = round(sum(dept_points) / len(dept_points), 1) if dept_points else None
        dept["minTotalPoints"] = min(dept_points) if dept_points else None
        dept["maxTotalPoints"] = max(dept_points) if dept_points else None


def load_current_lookup():
    data = json.loads(CURRENT_JSON.read_text(encoding="utf-8"))
    lookup = {}
    for school in data.get("schools", []):
        canonical = standardize_school_name(school.get("school", ""))
        lookup[canonical] = {
            "address": school.get("address", ""),
            "lat": school.get("lat"),
            "lng": school.get("lng"),
            "locationAccuracy": school.get("locationAccuracy", "missing"),
            "geocodeSource": school.get("geocodeSource", ""),
        }
    return lookup


def load_address_lookup():
    lookup = {}
    wb = load_workbook(ADDRESS_XLSX, read_only=False, data_only=True)
    ws = wb["學校彙總"]
    headers = [cell.value for cell in ws[1]]
    name_idx = headers.index("學校名稱")
    address_idx = headers.index("地址")
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = text(row[name_idx])
        address = text(row[address_idx])
        if not name or not address:
            continue
        lookup[standardize_school_name(name)] = address
    return lookup


def load_arcgis_lookup():
    if not ARCGIS_CACHE.exists():
        return {}
    raw = json.loads(ARCGIS_CACHE.read_text(encoding="utf-8"))
    return {address_key(key): value for key, value in raw.items()}


MANUAL_ADDRESS = {
    "嘉陽學校財團法人臺中市嘉陽高級中學": "臺中市清水區中航路3段1號",
    "嘉陽學校財團法人臺中市嘉陽高級中學進修部": "臺中市清水區中航路3段1號",
    "臺中市私立慈明高級中學": "臺中市太平區永隆里光德路388號",
    "臺中市私立慈明高級中學進修部": "臺中市太平區永隆里光德路388號",
}


def lookup_address_info(school_name, current_lookup, address_lookup):
    address_info = current_lookup.get(school_name, {})
    address = address_info.get("address") or address_lookup.get(school_name, "") or MANUAL_ADDRESS.get(school_name, "")
    if address:
        return address_info, address

    if school_name.endswith("進修部"):
        base_name = school_name.removesuffix("進修部")
        base_info = current_lookup.get(base_name, {})
        base_address = base_info.get("address") or address_lookup.get(base_name, "") or MANUAL_ADDRESS.get(base_name, "")
        if base_address:
            return base_info, base_address

    return address_info, address


def is_taichung_school(school_name, address):
    if "臺中" in address or "台中" in address:
        return True
    taichung_tokens = [
        "臺中",
        "台中",
        "中科實驗",
        "中興大學附屬臺中",
        "中興大學附屬高級中學",
        "光華學校財團法人臺中市",
        "嘉陽學校財團法人臺中市",
        "宜寧學校財團法人臺中市",
        "明德學校財團法人臺中市",
        "常春藤學校財團法人臺中市",
        "葳格學校財團法人臺中市",
        "東海大學附屬",
    ]
    return any(token in school_name for token in taichung_tokens)


def load_rows():
    wb = load_workbook(INPUT_XLSX, read_only=False, data_only=True)
    ws = wb["高中職最低三筆"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {header: values[idx] for idx, header in enumerate(headers)}
        rows.append(row)
    return rows


def build_site_payload():
    current_lookup = load_current_lookup()
    address_lookup = load_address_lookup()
    arcgis_lookup = load_arcgis_lookup()
    rows = load_rows()

    school_map = {}
    skipped = defaultdict(int)

    for row in rows:
        school_name = text(row["學校"])
        department = text(row["科系"])
        address_info, address = lookup_address_info(school_name, current_lookup, address_lookup)

        if not is_taichung_school(school_name, address):
            skipped["非台中學校"] += 1
            continue

        lat = address_info.get("lat")
        lng = address_info.get("lng")
        location_accuracy = address_info.get("locationAccuracy") or "missing"
        geocode_source = address_info.get("geocodeSource") or ""

        cached = arcgis_lookup.get(address_key(address))
        if (lat is None or lng is None) and cached:
            lat = cached.get("lat")
            lng = cached.get("lng")
            location_accuracy = "exact" if lat is not None and lng is not None else "missing"
            geocode_source = f"arcgis_address:{cached.get('score', '')}:{cached.get('matchedAddress', '')}".strip(":")

        if not address:
            skipped["缺地址"] += 1

        if school_name not in school_map:
            school_map[school_name] = {
                "school": school_name,
                "address": address,
                "sourceGroup": "main",
                "lat": lat,
                "lng": lng,
                "locationAccuracy": location_accuracy,
                "departments": [],
            }
            if geocode_source:
                school_map[school_name]["geocodeSource"] = geocode_source

        school = school_map[school_name]
        if not school.get("address") and address:
            school["address"] = address
        if school.get("lat") is None and lat is not None:
            school["lat"] = lat
            school["lng"] = lng
            school["locationAccuracy"] = location_accuracy
            if geocode_source:
                school["geocodeSource"] = geocode_source

        dept = next((item for item in school["departments"] if item["name"] == department), None)
        if dept is None:
            dept = {"name": department, "records": []}
            school["departments"].append(dept)

        subjects = {
            "國文": text(row["國文"]),
            "英語": text(row["英語"]),
            "數學": text(row["數學"]),
            "社會": text(row["社會"]),
            "自然": text(row["自然"]),
        }
        counts = {key: 0 for key in ["A", "B", "C"]}
        for grade in subjects.values():
            group = grade_group(grade)
            if group:
                counts[group] += 1

        dept["records"].append({
            "year": f"{int(row['年份'])}年",
            "category": "高中職",
            "school": school_name,
            "originalSchool": text(row.get("原始學校名稱")),
            "address": school.get("address", ""),
            "department": department,
            "subjects": subjects,
            "aCount": counts["A"],
            "bCount": counts["B"],
            "cCount": counts["C"],
            "writing": text(row["寫作"]),
            "score": row["積分"],
            "totalPoints": row["總積點"],
            "notes": text(row.get("備註")),
            "sourceGroup": "main",
        })

    schools = list(school_map.values())
    for school in schools:
        for dept in school["departments"]:
            dept["records"].sort(key=lambda record: (-int(record["year"].replace("年", "")), record["totalPoints"] or 999, record["score"] or 999))
        recalc_school(school)
        school["departments"].sort(key=lambda dept: (-(dept.get("avgTotalPoints") or 0), dept["name"]))
    schools.sort(key=lambda school: (-(school.get("avgTotalPoints") or 0), school["school"]))

    mapped = sum(1 for school in schools if school.get("lat") is not None and school.get("lng") is not None)
    records = sum(school["recordCount"] for school in schools)
    now = datetime.now(timezone(timedelta(hours=8))).replace(microsecond=0).isoformat()

    return {
        "generatedFrom": INPUT_XLSX.name,
        "summary": {
            "schools": len(schools),
            "mapped": mapped,
            "unmapped": len(schools) - mapped,
            "records": records,
        },
        "schools": schools,
        "importNotes": {
            "scope": "104-114年高中職最低三筆，僅匯入台中高中職學校。",
            "skipped": dict(skipped),
        },
        "updatedAt": now,
    }


def main():
    payload = build_site_payload()
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))
    print(json.dumps(payload["importNotes"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
