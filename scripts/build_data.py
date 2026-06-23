import json
import math
import time
import unicodedata
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Volumes/盛田協作/Ryuk/國三落點分析")
EXCEL = ROOT / "大華國中歷屆會考分數與免試分發學校對照表104-114_110-114落點分析.xlsx"
OUT_DIR = ROOT / "school-map" / "data"
CACHE = OUT_DIR / "geocode-cache.json"
OUT_JSON = OUT_DIR / "schools.json"

SOURCE_SHEETS = [
    ("最低錄取整理", 12, "main"),
]

DISTRICT_CENTERS = {
    "中區": (24.1432, 120.6814),
    "東區": (24.1375, 120.6970),
    "南區": (24.1209, 120.6660),
    "西區": (24.1416, 120.6629),
    "北區": (24.1587, 120.6838),
    "西屯區": (24.1810, 120.6270),
    "南屯區": (24.1384, 120.6164),
    "北屯區": (24.1840, 120.7015),
    "豐原區": (24.2521, 120.7225),
    "大里區": (24.0994, 120.6810),
    "太平區": (24.1266, 120.7180),
    "清水區": (24.2699, 120.5750),
    "沙鹿區": (24.2370, 120.5610),
    "大甲區": (24.3450, 120.6250),
    "東勢區": (24.2580, 120.8290),
    "霧峰區": (24.0610, 120.7000),
    "后里區": (24.3090, 120.7220),
    "新社區": (24.2430, 120.8090),
    "梧棲區": (24.2550, 120.5300),
    "龍井區": (24.2000, 120.5450),
    "神岡區": (24.2570, 120.6620),
    "潭子區": (24.2110, 120.7050),
    "烏日區": (24.1050, 120.6250),
    "大雅區": (24.2250, 120.6470),
}

COORDINATE_OVERRIDES = {
    "玉山高級中學": {
        "lat": 24.2517210,
        "lng": 120.8424720,
        "display": "玉山高中, 東崎路四段, 東勢區, 臺中市",
        "accuracy": "exact",
    },
    "私立玉山高級中學": {
        "lat": 24.2517210,
        "lng": 120.8424720,
        "display": "玉山高中, 東崎路四段, 東勢區, 臺中市",
        "accuracy": "exact",
    },
    "明德高級中學": {
        "lat": 24.1201528,
        "lng": 120.6829543,
        "display": "明德中學, 84, 明德街, 南區, 臺中市",
        "accuracy": "exact",
    },
    "私立明德高級中學": {
        "lat": 24.1201528,
        "lng": 120.6829543,
        "display": "明德中學, 84, 明德街, 南區, 臺中市",
        "accuracy": "exact",
    },
    "中港高級中學": {
        "lat": 24.2447691,
        "lng": 120.5429717,
        "display": "中港高中, 400, 文昌路, 梧棲區, 臺中市",
        "accuracy": "exact",
    },
}


def clean(value):
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def number(value):
    try:
        if value in ("", None):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def norm_address(address):
    text = clean(address)
    if not text:
        return ""
    text = text.replace(" ", "")
    if "臺中市" not in text and "台中市" not in text:
        text = f"臺中市{text}"
    return text


def approximate_location(address, school):
    for district, center in DISTRICT_CENTERS.items():
        if district in address:
            seed = sum(ord(ch) for ch in school)
            lat_offset = ((seed % 17) - 8) * 0.0018
            lng_offset = (((seed // 17) % 17) - 8) * 0.0018
            return {
                "lat": round(center[0] + lat_offset, 6),
                "lng": round(center[1] + lng_offset, 6),
                "display": f"{district} approximate",
                "accuracy": "approximate",
            }
    return {"lat": None, "lng": None, "error": "no_district_fallback", "accuracy": "missing"}


def geocode(school, address, cache):
    if not school and not address:
        return None
    for key, location in COORDINATE_OVERRIDES.items():
        if key in school:
            return location
    cache_key = f"school::{school}::{address}"
    if cache_key in cache:
        return cache[cache_key]

    attempts = [school]
    data = []
    error = None
    for attempt in attempts:
        query = urllib.parse.urlencode({
            "q": f"{attempt}, Taiwan",
            "format": "jsonv2",
            "limit": 1,
            "accept-language": "zh-TW",
        })
        url = f"https://nominatim.openstreetmap.org/search?{query}"
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": "dahua-admission-map/1.0 (local data visualization)",
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=8) as response:
                data = json.loads(response.read().decode("utf-8"))
        except Exception as exc:
            error = str(exc)
        if data:
            break
        time.sleep(0.4)

    if data:
        cache[cache_key] = {
            "lat": float(data[0]["lat"]),
            "lng": float(data[0]["lon"]),
            "display": data[0].get("display_name", ""),
            "accuracy": "exact",
        }
    else:
        cache[cache_key] = {"lat": None, "lng": None, "error": error or "not_found", "accuracy": "missing"}
    CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), "utf-8")
    time.sleep(0.8)
    return cache[cache_key]


def row_to_record(headers, values, source_group):
    item = dict(zip(headers, values))
    subjects = [clean(item.get(name)) for name in ["國文", "英語", "數學", "社會", "自然"]]
    return {
        "year": clean(item.get("年度")),
        "category": clean(item.get("類別")),
        "school": clean(item.get("錄取學校")),
        "address": clean(item.get("學校地址")),
        "department": clean(item.get("科組/科系")),
        "pathway": clean(item.get("管道名稱")),
        "subjects": {
            "國文": subjects[0],
            "英語": subjects[1],
            "數學": subjects[2],
            "社會": subjects[3],
            "自然": subjects[4],
        },
        "aCount": int(item.get("A數") or 0),
        "bCount": int(item.get("B數") or 0),
        "cCount": int(item.get("C數") or 0),
        "writing": clean(item.get("寫作")),
        "score": number(item.get("積分")),
        "totalPoints": number(item.get("總積點")),
        "notes": clean(item.get("備註")),
        "special": clean(item.get("特殊備註")),
        "reason": clean(item.get("篩選原因")),
        "sourceGroup": source_group,
    }


def mean(values):
    values = [value for value in values if value is not None]
    return round(sum(values) / len(values), 1) if values else None


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    cache = json.loads(CACHE.read_text("utf-8")) if CACHE.exists() else {}
    wb = load_workbook(EXCEL, data_only=True)

    records = []
    for sheet_name, header_row, source_group in SOURCE_SHEETS:
        ws = wb[sheet_name]
        headers = [clean(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
        for row in range(header_row + 1, ws.max_row + 1):
            values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
            if not values[0] or not values[2]:
                continue
            records.append(row_to_record(headers, values, source_group))

    grouped = defaultdict(list)
    for record in records:
        key = (record["school"], record["address"], record["sourceGroup"])
        grouped[key].append(record)

    schools = []
    for (school, address, source_group), rows in grouped.items():
        normalized_address = norm_address(address)
        location = geocode(school, normalized_address, cache) if normalized_address else None
        if location and (location.get("lat") is None or location.get("lng") is None):
            location = approximate_location(normalized_address, school)
        departments = defaultdict(list)
        for row in rows:
            departments[row["department"]].append(row)

        department_items = []
        for department, dept_rows in departments.items():
            department_items.append({
                "name": department,
                "avgTotalPoints": mean([row["totalPoints"] for row in dept_rows]),
                "minTotalPoints": min([row["totalPoints"] for row in dept_rows if row["totalPoints"] is not None], default=None),
                "maxTotalPoints": max([row["totalPoints"] for row in dept_rows if row["totalPoints"] is not None], default=None),
                "records": sorted(dept_rows, key=lambda row: (row["year"], row["totalPoints"] or -1), reverse=True),
            })

        all_points = [row["totalPoints"] for row in rows if row["totalPoints"] is not None]
        schools.append({
            "school": school,
            "address": normalized_address,
            "sourceGroup": source_group,
            "lat": location.get("lat") if location else None,
            "lng": location.get("lng") if location else None,
            "locationAccuracy": (location.get("accuracy") or "exact") if location and location.get("lat") is not None else "missing",
            "avgTotalPoints": mean(all_points),
            "minTotalPoints": min(all_points) if all_points else None,
            "maxTotalPoints": max(all_points) if all_points else None,
            "recordCount": len(rows),
            "departments": sorted(department_items, key=lambda dept: (dept["avgTotalPoints"] or -1, dept["name"]), reverse=True),
        })

    mapped = [school for school in schools if school["lat"] is not None and school["lng"] is not None]
    unmapped = [school for school in schools if school["lat"] is None or school["lng"] is None]

    payload = {
        "generatedFrom": EXCEL.name,
        "summary": {
            "schools": len(schools),
            "mapped": len(mapped),
            "unmapped": len(unmapped),
            "records": len(records),
        },
        "schools": sorted(schools, key=lambda item: (item["avgTotalPoints"] or -1, item["school"]), reverse=True),
    }

    CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), "utf-8")
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), "utf-8")
    print(json.dumps(payload["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
